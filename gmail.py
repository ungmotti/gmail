import imaplib, email, re, requests
import os, datetime, time
import exifread as ef
from urllib.parse import unquote
import unicodedata
import csv
import openpyxl
import hashlib
import gmplot

ORG_EMAIL   = "@gmail.com"
your_id = ''
FROM_EMAIL  = your_id + ORG_EMAIL
FROM_PWD    = ''
SMTP_SERVER = "imap.gmail.com"
SMTP_PORT   = 993



def get_Valid(valLst):
    # 메일 서버에 접속하여 이메일 내용을 읽음
    try:
        mail = imaplib.IMAP4_SSL(SMTP_SERVER)
        mail.login(FROM_EMAIL,FROM_PWD)
        mail.select('inbox')
        date = datetime.date.today().strftime("%d-%b-%Y")
        typ, data = mail.search(None, '(SENTSINCE {date} FROM "Kyle Choi")'.format(date=date))
        mail_ids = data[0]

        id_list = mail_ids.split()

        for i in id_list:
            typ, data = mail.fetch(i, '(RFC822)')

            for response_part in data:
                if isinstance(response_part, tuple):
                    try:
                        msg = email.message_from_string(response_part[1].decode('utf-8'))
                    except:
                        msg = email.message_from_string(response_part[1].decode('cp949'))

                    for part in msg.walk():

                        if part.get_content_type() == "text/plain":
                            try:
                                body = part.get_payload(decode=True).decode('utf-8')
                            except:
                                body = part.get_payload(decode=True).decode('cp949')
# 이메일에서 .kr/ , .ly/등의 패턴을 먼저 찾고, 먼저 왼쪽으로 전수조사하여 접속이 가능한 단축 url사이트를 찾음.
# 접속이 되는 단축 url 사이트를 찾으면 .kr/ 뒤로 전수조사를 통해 접속 가능한 url을 찾고, 해당 url의 full url에
# jpg, bmp, png가 있으면 그 url의 단축 url, full url, 해당 url이 포함되어 있던 메일의 수신 시각, 파일 이름을 리스트에 추가함.
                            m = newpat.search(body)
                            stemURL = ''
                            if m:
                                for i in range(3, len(m.group(2)) + 3):
                                    tryURL = 'http://' + m.group(2)[-i:] + m.group(3)
                                    try:
                                        req = requests.get(tryURL)
                                        if req.status_code == 200:
                                            stemURL = tryURL
                                            break

                                    except:
                                        continue

                                for i in range(1, 10):
                                    tryURL = stemURL + body[m.end(): m.end() + i]
                                    try:
                                        req = requests.get(tryURL)
                                        if req.status_code == 200:
                                            if '.jpg' in req.url or '.png' in req.url or '.bmp' in req.url:
                                                #print("Got it! %s" %tryURL)
                                                date = msg['Date']
                                                parsed_date = email.utils.parsedate(date)
                                                intTime = time.mktime((parsed_date))
                                                date = datetime.datetime.fromtimestamp(intTime) + datetime.timedelta(hours=9)
                                                date = date.strftime('%Y-%m-%d %H:%M:%S')
                                                valLst.append(date)
                                                valLst.append(tryURL)
                                                url = unquote(req.url)
                                                valLst.append(url)
                                                filename = url.split('/')[-1]
                                                valLst.append(filename.encode('utf-8').decode('utf-8'))

                                                break

                                            else:
                                                continue
                                        else:
                                            continue

                                    except:
                                        continue

                                continue

    except Exception as e:
        print(str(e))
    return valLst

# 실행 날짜를 폴더 이름으로 폴더 생성 후
# get_Valid 함수에서 반환받은 리스트로 엑셀파일 생성
def writeExcel(urlLst):
    if not os.path.exists(newpath):
        os.makedirs(newpath)
    xlsxpath = newpath + '/'+ todstr + '.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active


    for i in range(1, 5):
        j=1
        for k in range(i-1, len(urlLst), 4):
            if i == 4:
                new_name = unicodedata.normalize('NFC', urlLst[k])
                ws.cell(row=j, column = i+1).value = new_name
            ws.cell(row=j, column = i+1).value = urlLst[k]
            ws.cell(row=j, column = 1).value = j
            j += 1
    wb.save(xlsxpath)
    return xlsxpath

# writeExcel 함수에서 반환한 excel파일에서 full url을 알아내고 접속하여 사진을 다운로드 받음.
def getPhoto(xlsxpath):
    wb = openpyxl.load_workbook(xlsxpath)
    ws = wb['Sheet']
    i = 1
    while True:
        full_url = ws.cell(row = i, column = 4).value
        filename = ws.cell(row = i, column = 5).value
        if full_url == None:
            break
        data = requests.get(full_url)
        with open(newpath + '/' + filename, 'wb') as f:
            f.write(data.content)
        i += 1

# exifread 모듈로 파싱한 gps 정보를 구글 포맷에 맞게 변환하는 함수
def _convert_to_degress(value):
    d = float(value.values[0].num) / float(value.values[0].den)
    m = float(value.values[1].num) / float(value.values[1].den)
    s = float(value.values[2].num) / float(value.values[2].den)

    return d + (m / 60.0) + (s / 3600.0)

# GPS와 해시값 산출하고 엑셀 파일에 기록
def GPSandHash(xlsxpath):
    wb = openpyxl.load_workbook(xlsxpath)
    ws = wb['Sheet']
    i = 1
    while True:
        filename = ws.cell(row = i, column = 5).value
        if filename == None:
            break
        filename = newpath + '/' + filename
        with open (filename, 'rb') as f:
            tags = ef.process_file(f)
            latitude = tags.get('GPS GPSLatitude')
            latitude_ref = tags.get('GPS GPSLatitudeRef')
            longitude = tags.get('GPS GPSLongitude')
            longitude_ref = tags.get('GPS GPSLongitudeRef')
            if latitude:
                lat_value = _convert_to_degress(latitude)
                if latitude_ref.values != 'N':
                    lat_value = -lat_value
                ws.cell(row=i, column=6).value = lat_value
            else:
                ws.cell(row=i, column=6).value = 'N/A'


            if longitude:
                lon_value = _convert_to_degress(longitude)
                if longitude_ref.values != 'E':
                    lon_value = -lon_value
                ws.cell(row=i, column=7).value = lon_value
            else:
                ws.cell(row=i, column=7).value = 'N/A'


            f.seek(0)
            BlockSize = 512
            hash_md5 = hashlib.md5()
            hash_sha1 = hashlib.sha1()
            buf =  f.read(BlockSize)


            while len(buf) > 0:
                hash_md5.update(buf)
                hash_sha1.update(buf)
                buf = f.read(BlockSize)
            ws.cell(row = i, column = 8).value = hash_md5.hexdigest()
            ws.cell(row = i, column = 9).value = hash_sha1.hexdigest()
        i+=1

    wb.template = False
    wb.save(xlsxpath)

# 엑셀 파일에 기록된 GPS정보 기반으로 지도에 GPS Marker찍고 연결하기(해당 날짜 폴더).
def GPSmarker(xlsxpath):
    wb = openpyxl.load_workbook(xlsxpath)
    ws = wb['Sheet']
    i = 1
    latLst=[]
    lonLst=[]
    DateLst=[]
    while True:

        Date = ws.cell(row = i, column = 2).value
        lat = ws.cell(row = i,  column = 6).value
        lon = ws.cell(row = i,  column = 7).value
        if Date == None:
            break

        if lat == 'N/A':
            i += 1
            continue
        else:
            DateLst.append(Date)
            latLst.append(lat)
            lonLst.append(lon)

        i += 1

    if latLst != []:
        gmap = gmplot.GoogleMapPlotter(latLst[0], lonLst[0], 8)
        for i in range(len(latLst)):
            gmap.marker(latLst[i], lonLst[i], 'red', title=DateLst[i])
        gmap.plot(latLst, lonLst, 'cornflowerblue', edge_width=7)

        gmap.draw(newpath + '/map1.html')


# 엑셀 파일을 읽어서 CSV파일로 작성하기(해당 날짜 폴더)
def csvWriter(xlsxpath):
    wb = openpyxl.load_workbook(xlsxpath)
    ws = wb['Sheet']
    title = ['Number', 'Date', 'Shortened URL', 'Full URL', 'File Name','Latitude', 'Longitude', 'MD5', 'SHA1']

    with open(newpath + '/' + todstr + '.csv', 'w', newline="", encoding='utf-8') as f:  # open('test.csv', 'w', newline="") for python 3
        c = csv.writer(f)
        c.writerow(title)
        for r in ws.rows:
            c.writerow([cell.value for cell in r])

# 엑셀 파일을 읽어서 통합 DB에 추가적으로 기록, 통합 지도에 Marker 찍기
def dbWriter(xlsxpath):
    wb = openpyxl.load_workbook(xlsxpath)
    ws = wb['Sheet']

    with open(rootpath + '/gmailDB.csv', 'a', newline="", encoding='utf-8') as f, open(rootpath + '/gmailDB.csv', 'r',  encoding='utf-8') as rf:
        c = csv.writer(f)
        cr = csv.reader(rf)
        linenum=0
        for row in cr:
            linenum = cr.line_num
        for i, r in enumerate(ws.rows):
            if r[0].value != None:
                c.writerow([linenum+i+1] + [cell.value for cell in r[1:]])
    DateLst = []
    latLst = []
    lonLst = []
    with open(rootpath + '/gmailDB.csv', 'r', encoding='utf-8') as f:
        c = csv.reader(f)
        for line in c:
            if line != []:
                try:
                    if float(line[5]) > 0 or float(line[6]) <= 0:
                        DateLst.append(line[1])
                        latLst.append(float(line[5]))
                        lonLst.append(float(line[6]))
                except:
                    continue
            else:
                break
    gmap = gmplot.GoogleMapPlotter(latLst[0], lonLst[0], 8)
    for i in range(len(latLst)):
        gmap.marker(latLst[i], lonLst[i], 'red', title=DateLst[i])
    gmap.plot(latLst, lonLst, 'cornflowerblue', edge_width=7)
    gmap.draw(rootpath+ '/map1.html')

tod = datetime.date.today()
todstr = tod.isoformat()
rootpath = r''
newpath = rootpath + todstr
newpat = re.compile("(http.?://)?(.*?)(\.[a-z]{2,3}/)")
urlLst=[]
invalLst=[]
ValidLst= get_Valid(urlLst)
xlsxpath = writeExcel(ValidLst)
getPhoto(xlsxpath)
GPSandHash(xlsxpath)
GPSmarker(xlsxpath)
csvWriter(xlsxpath)
dbWriter(xlsxpath)
